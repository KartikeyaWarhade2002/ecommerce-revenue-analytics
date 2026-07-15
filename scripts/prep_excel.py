from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =====================================================
# Paths
# =====================================================

INPUT_FILE = Path("data/processed/analytics_orders.csv")
OUTPUT_FILE = Path("excel/analysis.xlsx")

# =====================================================
# Read Dataset
# =====================================================

df = pd.read_csv(
    INPUT_FILE,
    parse_dates=["order_purchase_timestamp"]
)

# =====================================================
# Create Date Columns
# =====================================================

df["Order Year"] = df["order_purchase_timestamp"].dt.year
df["Order Month"] = df["order_purchase_timestamp"].dt.strftime("%Y-%m")

# =====================================================
# Create Revenue Column
# Revenue = Product Price (Project Standard)
# =====================================================

df["Revenue"] = df["price"]

# =====================================================
# Rename Columns
# =====================================================

df.rename(
    columns={
        "order_id": "Order ID",
        "customer_unique_id": "Customer ID",
        "customer_city": "City",
        "customer_state": "State",
        "product_category": "Category",
        "payment_type": "Payment Type",
        "order_status": "Order Status",
        "freight_value": "Freight Cost",
    },
    inplace=True,
)

# =====================================================
# Select Excel Columns
# =====================================================

excel_df = df[
    [
        "Order ID",
        "Order Year",
        "Order Month",
        "Customer ID",
        "City",
        "State",
        "Category",
        "Revenue",
        "Freight Cost",
        "Payment Type",
        "Order Status",
    ]
]

# =====================================================
# Create Output Folder
# =====================================================

OUTPUT_FILE.parent.mkdir(
    parents=True,
    exist_ok=True
)

# =====================================================
# Save Excel Workbook
# =====================================================

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    excel_df.to_excel(
        writer,
        sheet_name="Data",
        index=False
    )

# =====================================================
# Formatting Workbook
# =====================================================

workbook = load_workbook(OUTPUT_FILE)
worksheet = workbook["Data"]

# Freeze Header
worksheet.freeze_panes = "A2"

# Auto-fit Columns
for column_cells in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(column_cells[0].column)

    for cell in column_cells:
        try:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        except Exception:
            pass

    worksheet.column_dimensions[column_letter].width = max_length + 3

workbook.save(OUTPUT_FILE)

# =====================================================
# Summary
# =====================================================

print("=" * 70)
print("Excel dataset created successfully.")
print("=" * 70)
print(f"Rows    : {len(excel_df):,}")
print(f"Columns : {len(excel_df.columns)}")
print(f"Saved to: {OUTPUT_FILE}")
print("=" * 70)